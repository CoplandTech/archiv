import os
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
from aiogram import types
from aiogram.dispatcher.filters import Text
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup, Message

from loader import dp
from module.TinyDB.config import Query, db_users
from module.keyboards import keyboard_admin, keyboard_start
from module.utils import AdminFilter
from config import PHRASES, TEMP_DIR

@dp.message_handler(commands=['admin'], is_admin=True)
async def admin(message: types.Message):
    await message.answer(text=PHRASES["btn_login"], reply_markup=keyboard_admin)

async def generate_xlsx(filename, headers, rows):
    filepath = os.path.join(TEMP_DIR, filename)
    os.makedirs(TEMP_DIR, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active

    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row_num, row in enumerate(rows, start=2):
        for col_num, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left")

    for col_num, width in zip(range(1, len(headers) + 1), [3, 11, 15, 15]):
        sheet.column_dimensions[sheet.cell(row=1, column=col_num).column_letter].width = width

    # Сохраняем файл
    workbook.save(filepath)
    return filepath

@dp.message_handler(Text(equals=PHRASES["btn_users_admin"]), is_admin=True)
async def users_admin_menu(message: Message):
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.row(
        InlineKeyboardButton("ID, имена и телефоны", callback_data="export_id_names_phones"),
        InlineKeyboardButton("Номера и купоны", callback_data="export_numbers_coupons")
    )
    keyboard.add(
        InlineKeyboardButton("Выгрузить всю базу", callback_data="export_full_db")
    )

    await message.answer("Выберите выгрузку:", reply_markup=keyboard)

@dp.callback_query_handler(Text(equals="export_id_names_phones"), is_admin=True)
async def export_id_names_phones(callback_query):
    await callback_query.answer()
    headers = ["#", "ID", "Full Name", "Phone Number"]
    rows = []

    for i, user in enumerate(db_users.all(), start=1):
        rows.append([
            i,
            user.get("id", "Не указано"),
            user.get("full_name", "Не указано"),
            user.get("phone_number", "Не указано")
        ])

    filepath = await generate_xlsx("id_names_phones.xlsx", headers, rows)
    await callback_query.message.answer_document(open(filepath, "rb"))
    try:
        os.remove(filepath)
    except FileNotFoundError:
        pass

@dp.callback_query_handler(Text(equals="export_numbers_coupons"), is_admin=True)
async def export_numbers_coupons(callback_query):
    await callback_query.answer()
    headers = ["#", "Phone Number", "Coupon Code"]
    rows = []

    for i, user in enumerate(db_users.all(), start=1):
        rows.append([
            i,
            user.get("phone_number", "Не указано"),
            user.get("CouponCode", "Не указано")
        ])

    filepath = await generate_xlsx("numbers_coupons.xlsx", headers, rows)
    await callback_query.message.answer_document(open(filepath, "rb"))
    try:
        os.remove(filepath)
    except FileNotFoundError:
        pass

@dp.callback_query_handler(Text(equals="export_full_db"), is_admin=True)
async def export_full_db(callback_query):
    await callback_query.answer()
    headers = ["#"] + (list(db_users.all()[0].keys()) if db_users.all() else [])
    rows = []

    for i, user in enumerate(db_users.all(), start=1):
        row = [i]
        for key in headers[1:]:
            value = user.get(key, "Не указано")
            if isinstance(value, str) and "T" in value:
                try:
                    value = datetime.fromisoformat(value).strftime("%d.%m.%Y %H:%M")
                except ValueError:
                    pass
            row.append(value)
        rows.append(row)

    filepath = await generate_xlsx("full_db.xlsx", headers, rows)
    await callback_query.message.answer_document(open(filepath, "rb"))
    try:
        os.remove(filepath)
    except FileNotFoundError:
        pass

@dp.message_handler(Text(equals=PHRASES["btn_back"]))
async def admin(message: types.Message):
    await message.answer(text=PHRASES["btn_logout"], reply_markup=keyboard_start)
